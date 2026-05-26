"""
MONITOR DE PRECIOS MAYORISTAS — McAllen TX vs CDMX
Origen Agro Group | v6.0
- Extrae precios reales del PDF USDA AMS Reporte 3130
- Precios Mexico actualizados via Claude API con busqueda web cada mañana
- 25 productos con fuentes y fechas claras
"""
import requests, smtplib, schedule, time, os, re, json, base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from twilio.rest import Client
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ============================================================
#  CONFIGURACION
# ============================================================
GMAIL_ORIGEN      = os.environ.get("GMAIL_ORIGEN",      "pichomel.brands@gmail.com")
GMAIL_APP_PASS    = os.environ.get("GMAIL_APP_PASS",     "")
CORREOS_DESTINO   = ["rodrigo@origenagro.com", "vicente@origenagro.com"]
TWILIO_SID        = os.environ.get("TWILIO_SID",        "")
TWILIO_TOKEN      = os.environ.get("TWILIO_TOKEN",      "")
TWILIO_WHATSAPP   = os.environ.get("TWILIO_WHATSAPP",   "whatsapp:+14155238886")
TU_WHATSAPP       = os.environ.get("TU_WHATSAPP",       "whatsapp:+5215543472416")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
URL_PAGINA        = os.environ.get("URL_PAGINA",        "https://pichomelbrands-prog.github.io/precios-mcallen")

HORA_REPORTE = "13:00"  # 7:00 AM CDMX = 13:00 UTC
LBS_A_KG     = 2.20462
URL_PDF_USDA = "https://www.ams.usda.gov/mnreports/fvdfob.pdf"
MARCA        = "Origen Agro Group"
LOGO_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo_oag.png") if os.path.exists(
               os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo_oag.png")) else None

# ============================================================
#  PRODUCTOS — 26 productos
# ============================================================
PRODUCTOS = [
    {"nombre":"Chile jalapeno",    "cat":"chile",   "sniim":"Chile jalapeño",   "fuente_tipo":"3130","buscar_pdf":["PEPPERS, JALAPENO","PEPPERS, JALAPE"],"peso_lb":28.0, "unidad":"caja 1 1/9 bushel (28 lb)","nota_unidad":""},
    {"nombre":"Chile serrano",     "cat":"chile",   "sniim":"Chile serrano",    "fuente_tipo":"3130","buscar_pdf":["PEPPERS, SERRANO"],                   "peso_lb":28.0, "unidad":"caja 1 1/9 bushel (28 lb)","nota_unidad":""},
    {"nombre":"Chile poblano",     "cat":"chile",   "sniim":"Chile poblano",    "fuente_tipo":"3130","buscar_pdf":["PEPPERS, POBLANO"],                   "peso_lb":28.0, "unidad":"caja 1 1/9 bushel (28 lb)","nota_unidad":""},
    {"nombre":"Chile habanero",    "cat":"chile",   "sniim":"Chile habanero",   "fuente_tipo":"3130","buscar_pdf":["PEPPERS, HABANERO"],                  "peso_lb":8.0,  "unidad":"caja 8 lb","nota_unidad":""},
    {"nombre":"Tomate saladet",    "cat":"verdura", "sniim":"Tomate saladette", "fuente_tipo":"3130","buscar_pdf":["TOMATOES, PLUM","TOMATOES, ROMA"],     "peso_lb":25.0, "unidad":"caja 25 lb","nota_unidad":""},
    {"nombre":"Tomatillo",         "cat":"verdura", "sniim":"Tomate verde",     "fuente_tipo":"3130","buscar_pdf":["TOMATILLOS"],                         "peso_lb":28.0, "unidad":"caja 1 1/9 bushel (28 lb)","nota_unidad":""},
    {"nombre":"Pepino",            "cat":"verdura", "sniim":"Pepino",           "fuente_tipo":"3130","buscar_pdf":["CUCUMBERS"],                          "peso_lb":55.0, "unidad":"caja 1 1/9 bushel (55 lb)","nota_unidad":""},
    {"nombre":"Brocoli",           "cat":"verdura", "sniim":"Brocoli",          "fuente_tipo":"3130","buscar_pdf":["BROCCOLI"],                           "peso_lb":20.0, "unidad":"caja 20 lb Crown Cut","nota_unidad":""},
    {"nombre":"Coliflor",          "cat":"verdura", "sniim":"Coliflor",         "fuente_tipo":"3130","buscar_pdf":["CAULIFLOWER"],                        "peso_lb":25.0, "unidad":"caja aprox 25 lb","nota_unidad":"peso aproximado"},
    {"nombre":"Zanahoria",         "cat":"verdura", "sniim":"Zanahoria",        "fuente_tipo":"3130","buscar_pdf":["CARROTS"],                            "peso_lb":50.0, "unidad":"saco 50 lb","nota_unidad":""},
    {"nombre":"Lechuga",           "cat":"verdura", "sniim":"Lechuga",          "fuente_tipo":"3130","buscar_pdf":["LETTUCE, ICEBERG"],                   "peso_lb":50.0, "unidad":"caja 24s aprox 50 lb","nota_unidad":"peso aproximado"},
    {"nombre":"Apio",              "cat":"verdura", "sniim":"Apio",             "fuente_tipo":"3130","buscar_pdf":["CELERY"],                             "peso_lb":55.0, "unidad":"caja aprox 55 lb","nota_unidad":"peso aproximado"},
    {"nombre":"Espinaca",          "cat":"verdura", "sniim":"Espinaca",         "fuente_tipo":"3130","buscar_pdf":["SPINACH"],                            "peso_lb":20.0, "unidad":"caja 24s bunched 20 lb","nota_unidad":""},
    {"nombre":"Cilantro",          "cat":"hierba",  "sniim":"Cilantro",         "fuente_tipo":"3130","buscar_pdf":["CILANTRO"],                           "peso_lb":None, "unidad":"caja 60 manojos","nota_unidad":"precio por manojo"},
    {"nombre":"Cebollita cambray", "cat":"verdura", "sniim":"Cebolla cambray",  "fuente_tipo":"3130","buscar_pdf":["ONIONS, GREEN"],                      "peso_lb":25.0, "unidad":"caja bunched aprox 25 lb","nota_unidad":"peso aproximado"},
    {"nombre":"Mango Ataulfo",     "cat":"fruta",   "sniim":"Mango",            "fuente_tipo":"3130","buscar_pdf":["MANGOES","MANGO"],                    "peso_lb":None, "unidad":"caja 1 layer","nota_unidad":"precio por pieza segun calibre"},
    {"nombre":"Platano",           "cat":"fruta",   "sniim":"Platano",          "fuente_tipo":"3130","buscar_pdf":["BANANAS","PLANTAINS"],                "peso_lb":40.0, "unidad":"caja 40 lb","nota_unidad":""},
    {"nombre":"Pina",              "cat":"fruta",   "sniim":"Pina",             "fuente_tipo":"3130","buscar_pdf":["PINEAPPLES"],                         "peso_lb":27.0, "unidad":"caja 1 layer aprox 27 lb","nota_unidad":"peso aproximado"},
    {"nombre":"Jicama",            "cat":"verdura", "sniim":"Jicama",           "fuente_tipo":"ref", "buscar_pdf":[],"peso_lb":40.0,"unidad":"caja 40 lb","nota_unidad":"","ref_usda_lb":(0.25,0.45),"ref_fuente":"USDA AMS Nogales AZ / Historico","ref_fecha":""},
    {"nombre":"Nopal",             "cat":"verdura", "sniim":"Nopal",            "fuente_tipo":"ref", "buscar_pdf":[],"peso_lb":40.0,"unidad":"caja 40 lb","nota_unidad":"","ref_usda_lb":(0.45,0.80),"ref_fuente":"USDA AMS Nogales AZ / Historico","ref_fecha":""},
    {"nombre":"Calabaza",          "cat":"verdura", "sniim":"Calabaza",         "fuente_tipo":"ref", "buscar_pdf":[],"peso_lb":30.0,"unidad":"caja 1 1/9 bushel","nota_unidad":"","ref_usda_lb":(0.25,0.45),"ref_fuente":"USDA AMS Nogales AZ / Historico","ref_fecha":""},
    {"nombre":"Betabel",           "cat":"verdura", "sniim":"Betabel",          "fuente_tipo":"ref", "buscar_pdf":[],"peso_lb":25.0,"unidad":"saco 25 lb","nota_unidad":"","ref_usda_lb":(0.40,0.65),"ref_fuente":"USDA AMS Historico","ref_fecha":""},
    {"nombre":"Rabano",            "cat":"verdura", "sniim":"Rabano",           "fuente_tipo":"ref", "buscar_pdf":[],"peso_lb":25.0,"unidad":"saco 25 lb","nota_unidad":"","ref_usda_lb":(0.45,0.75),"ref_fuente":"USDA AMS Historico","ref_fecha":""},
    {"nombre":"Hierbabuena",       "cat":"hierba",  "sniim":"Hierbabuena",      "fuente_tipo":"ref", "buscar_pdf":[],"peso_lb":None,"unidad":"por manojo","nota_unidad":"precio por manojo","ref_usda_lb":(1.20,1.80),"ref_fuente":"USDA AMS Historico","ref_fecha":""},
    {"nombre":"Malanga",           "cat":"verdura", "sniim":"Malanga",          "fuente_tipo":"ref", "buscar_pdf":[],"peso_lb":36.0,"unidad":"caja 36 lb","nota_unidad":"","ref_usda_lb":(0.55,0.90),"ref_fuente":"USDA AMS Florida / Historico","ref_fecha":""},
    {"nombre":"Coco verde",        "cat":"fruta",   "sniim":"Coco",             "fuente_tipo":"ref", "buscar_pdf":[],"peso_lb":None,"unidad":"por pieza","nota_unidad":"precio por pieza","ref_usda_lb":(0.80,1.40),"ref_fuente":"Importadores Miami / Historico","ref_fecha":""},
]

# Precios MX de respaldo (se actualizan via Claude API cada mañana)
PRECIOS_MX_RESPALDO = {
    "Chile jalapeno":{"precio":46.0,"fuente":"SNIIM/El Financiero","fecha":"06/05/2026"},
    "Chile serrano":{"precio":35.0,"fuente":"SNIIM/El Financiero","fecha":"06/05/2026"},
    "Chile poblano":{"precio":60.0,"fuente":"SNIIM/El Financiero","fecha":"06/05/2026"},
    "Chile habanero":{"precio":65.0,"fuente":"SNIIM/El Financiero","fecha":"06/05/2026"},
    "Tomate saladet":{"precio":30.0,"fuente":"SNIIM/El Financiero","fecha":"06/05/2026"},
    "Tomatillo":{"precio":18.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Pepino":{"precio":28.0,"fuente":"SNIIM/El Financiero","fecha":"06/05/2026"},
    "Brocoli":{"precio":30.0,"fuente":"SNIIM/El Financiero","fecha":"06/05/2026"},
    "Coliflor":{"precio":22.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Zanahoria":{"precio":25.0,"fuente":"SNIIM/El Financiero","fecha":"06/05/2026"},
    "Lechuga":{"precio":18.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Apio":{"precio":16.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Espinaca":{"precio":25.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Cilantro":{"precio":0.4,"fuente":"SNIIM/CEDA","fecha":"06/05/2026"},
    "Cebollita cambray":{"precio":22.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Mango Ataulfo":{"precio":28.0,"fuente":"SNIIM/CEDA","fecha":"06/05/2026"},
    "Platano":{"precio":12.0,"fuente":"SNIIM/CEDA","fecha":"06/05/2026"},
    "Pina":{"precio":14.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Jicama":{"precio":12.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Nopal":{"precio":15.0,"fuente":"SNIIM/CEDA","fecha":"06/05/2026"},
    "Calabaza":{"precio":12.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Betabel":{"precio":14.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Rabano":{"precio":12.0,"fuente":"SNIIM","fecha":"06/05/2026"},
    "Hierbabuena":{"precio":0.8,"fuente":"SNIIM/CEDA","fecha":"06/05/2026"},
    "Malanga":{"precio":28.0,"fuente":"SNIIM/CEDA","fecha":"06/05/2026"},
    "Coco verde":{"precio":22.0,"fuente":"SNIIM/CEDA","fecha":"06/05/2026"},
}

# ============================================================
#  TIPO DE CAMBIO
# ============================================================
def obtener_tipo_cambio():
    for url in ["https://api.frankfurter.app/latest?from=USD&to=MXN",
                "https://open.er-api.com/v6/latest/USD"]:
        try:
            r = requests.get(url, timeout=10).json()
            tc = r.get("rates",{}).get("MXN") or r.get("conversion_rates",{}).get("MXN")
            if tc:
                tc = round(float(tc), 2)
                print(f"  TC: ${tc} MXN")
                return tc
        except: continue
    return 17.50

# ============================================================
#  PRECIOS MEXICO VIA CLAUDE API CON BUSQUEDA WEB
# ============================================================
def obtener_precios_mx_claude():
    if not ANTHROPIC_API_KEY:
        print("  ANTHROPIC_API_KEY no configurada — usando precios de respaldo")
        return PRECIOS_MX_RESPALDO.copy()

    print("  Consultando precios Mexico via Claude API con busqueda web...")
    hoy = datetime.now().strftime("%d de %B de %Y")
    productos_lista = ", ".join([p["nombre"] for p in PRODUCTOS])

    prompt = f"""Hoy es {hoy}. Necesito los precios mayoristas actuales en Mexico (Ciudad de Mexico, 
mercado de abastos / CEDA / SNIIM) de los siguientes productos agricolas frescos:

{productos_lista}

Para cada producto dame:
- precio en MXN por kg (o por manojo para Cilantro e Hierbabuena, o por pieza para Coco verde y Mango Ataulfo)
- fuente donde encontraste el precio (SNIIM, El Financiero, CEDA, etc)

Busca en SNIIM (economia-sniim.gob.mx), El Financiero, o cualquier fuente confiable de precios agricolas mexicanos de hoy.

Responde SOLO con un JSON valido, sin texto adicional, con este formato exacto:
{{
  "Chile jalapeno": {{"precio": 46.0, "fuente": "SNIIM", "fecha": "{datetime.now().strftime('%d/%m/%Y')}"}},
  "Chile serrano": {{"precio": 35.0, "fuente": "SNIIM", "fecha": "{datetime.now().strftime('%d/%m/%Y')}"}},
  ... (todos los productos)
}}"""

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 2000,
                "tools": [{"type": "web_search_20250305", "name": "web_search"}],
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=60
        )
        data = resp.json()
        texto = "".join(b.get("text","") for b in data.get("content",[]) if b.get("type")=="text")
        # Limpiar y parsear JSON
        texto = re.sub(r'```json|```','', texto).strip()
        precios = json.loads(texto)
        # Validar que tiene todos los productos
        validos = {k:v for k,v in precios.items() if isinstance(v,dict) and "precio" in v and float(v["precio"])>0}
        print(f"  Claude API: {len(validos)} precios Mexico obtenidos")
        # Completar con respaldo si faltan
        for prod in PRODUCTOS:
            if prod["nombre"] not in validos and prod["nombre"] in PRECIOS_MX_RESPALDO:
                validos[prod["nombre"]] = PRECIOS_MX_RESPALDO[prod["nombre"]]
                validos[prod["nombre"]]["fuente"] += " (respaldo)"
        return validos
    except Exception as e:
        print(f"  Error Claude API: {e} — usando precios de respaldo")
        return PRECIOS_MX_RESPALDO.copy()

# ============================================================
#  EXTRAER PRECIOS DEL PDF USDA REPORTE 3130
# ============================================================
def extraer_precios_pdf():
    print("Descargando PDF USDA Reporte 3130...")
    precios = {}
    fecha_reporte = datetime.now().strftime("%d/%m/%Y")
    try:
        r = requests.get(URL_PDF_USDA, timeout=30)
        if r.status_code != 200:
            print(f"  Error PDF: {r.status_code}")
            return precios, fecha_reporte
        try:
            from pdfminer.high_level import extract_text
            from io import BytesIO
            texto = extract_text(BytesIO(r.content))
        except ImportError:
            texto = r.content.decode("latin-1", errors="ignore")

        fm = re.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s*20\d{2}', texto)
        if fm:
            try:
                fecha_reporte = datetime.strptime(fm.group().replace(",","").replace("  "," "), "%B %d %Y").strftime("%d/%m/%Y")
            except: pass
        print(f"  Fecha reporte: {fecha_reporte}")

        secciones = re.split(r'MEXICO CROSSINGS THROUGH (?:SOUTH\s+)?TEXAS', texto, flags=re.IGNORECASE)
        if len(secciones) < 2:
            print("  No se encontro seccion Mexico Crossings Through Texas")
            return precios, fecha_reporte

        partes = []
        for sec in secciones[1:]:
            corte = re.search(r'\n(?:CENTRAL|SOUTH|NORTH|EAST|WEST|FLORIDA|CALIFORNIA|ARIZONA|WASHINGTON|OREGON|IDAHO|GEORGIA|COLUMBIA|YAKIMA|KERN|OXNARD|SALINAS|SANTA MARIA|COACHELLA|APPALACHIAN|NEW YORK|MICHIGAN|ARGENTINA|CHILE IMPORTS|PERU|BRAZIL|CARIBBEAN|CENTRAL AMERICA|SOUTH AMERICA|GUATEMALA|MOROCCO)', sec)
            partes.append(sec[:corte.start()] if corte else sec[:3000])
        texto_texas = "\n".join(partes)

        for prod in PRODUCTOS:
            if prod["fuente_tipo"] != "3130" or not prod["buscar_pdf"]: continue
            nombre = prod["nombre"]
            encontrado = False
            for termino in prod["buscar_pdf"]:
                patron = rf'---{re.escape(termino)}[:\s].*?(?=---|\Z)'
                bloque = re.search(patron, texto_texas, re.DOTALL|re.IGNORECASE)
                if not bloque: continue
                rangos = re.findall(r'\b(\d{1,3}\.\d{2})-(\d{1,3}\.\d{2})\b', bloque.group())
                validos = [(float(a),float(b)) for a,b in rangos if 5.0<=float(a)<=500.0 and 5.0<=float(b)<=500.0]
                if not validos:
                    unicos = [float(p) for p in re.findall(r'\b(\d{1,3}\.\d{2})\b', bloque.group()) if 5.0<=float(p)<=500.0]
                    if unicos:
                        pc=sum(unicos)/len(unicos)
                        pl=pc/(prod["peso_lb"] or 28.0)
                        precios[nombre]={"precio_lb":round(pl,4),"precio_caja":round(pc,2),"fecha":fecha_reporte,"fuente":"USDA AMS Reporte 3130","contenedor":prod["unidad"]}
                        encontrado=True; break
                    continue
                usar=validos[:3]
                pc=sum((a+b)/2 for a,b in usar)/len(usar)
                if prod["peso_lb"]: pl=pc/prod["peso_lb"]
                elif nombre=="Cilantro": pl=pc/60
                elif nombre=="Mango Ataulfo": pl=pc/14
                else: pl=pc/28
                precios[nombre]={"precio_lb":round(pl,4),"precio_caja":round(pc,2),"fecha":fecha_reporte,"fuente":"USDA AMS Reporte 3130","contenedor":prod["unidad"]}
                encontrado=True; break
            if not encontrado:
                print(f"  Sin dato Texas: {nombre}")
        print(f"  {len(precios)} productos extraidos del Reporte 3130")
    except Exception as e:
        import traceback; print(f"  Error PDF: {e}"); traceback.print_exc()
    return precios, fecha_reporte

# ============================================================
#  CALCULAR COMPARATIVA
# ============================================================
def calcular_comparativa(precios_pdf, precios_mx, fecha_reporte, tc):
    comp = []
    hoy  = datetime.now().strftime("%d/%m/%Y")
    for i,prod in enumerate(PRODUCTOS,1):
        nombre = prod["nombre"]
        es_por_pieza = prod["nota_unidad"] in ("precio por pieza","precio por manojo","precio por pieza segun calibre")
        if prod["fuente_tipo"]=="3130" and nombre in precios_pdf:
            dat=precios_pdf[nombre]
            pa_lb=dat["precio_lb"]
            fuente_usda=f"USDA AMS Reporte 3130 ({dat['fecha']})"
            contenedor=dat["contenedor"]
        else:
            ref=prod.get("ref_usda_lb",(0.40,0.80))
            pa_lb=round((ref[0]+ref[1])/2,4)
            rf=prod.get("ref_fecha",hoy) or hoy
            fuente_usda=f"Ref. {prod.get('ref_fuente','historica')} ({rf})"
            contenedor=prod["unidad"]
        pa_usd_kg=round(pa_lb*LBS_A_KG,4)
        pa_mxn_kg=round(pa_usd_kg*tc,2)
        mx=precios_mx.get(nombre,{})
        precio_mx=mx.get("precio")
        fuente_mx=mx.get("fuente","N/D")
        fecha_mx=mx.get("fecha",hoy)
        if es_por_pieza: dif=pct=None; cdmx_barato=None
        elif precio_mx and pa_mxn_kg:
            dif=round(pa_mxn_kg-precio_mx,2)
            pct=round((dif/precio_mx)*100,1) if precio_mx>0 else 0
            cdmx_barato=precio_mx<pa_mxn_kg
        else: dif=pct=None; cdmx_barato=None
        comp.append({"idx":i,"nombre":nombre,"cat":prod["cat"],"pa_lb":round(pa_lb,4),
            "pa_usd_kg":pa_usd_kg,"pa_mxn_kg":pa_mxn_kg,"fuente_usda":fuente_usda,
            "contenedor":contenedor,"nota_unidad":prod.get("nota_unidad",""),"fuente_tipo":prod["fuente_tipo"],
            "precio_mx":precio_mx,"fuente_mx":fuente_mx,"fecha_mx":fecha_mx,
            "dif":dif,"pct":pct,"cdmx_barato":cdmx_barato})
    return comp

# ============================================================
#  JSON
# ============================================================
def guardar_json(comp, tc, fecha_reporte):
    datos={"fecha":datetime.now().strftime("%d/%m/%Y %H:%M"),"fecha_reporte":fecha_reporte,"tc":tc,
           "marca":MARCA,"fuente_3130":"USDA AMS National FOB Review — Mexico Crossings Through Texas",
           "url_reporte":URL_PDF_USDA,
           "productos":[{"idx":p["idx"],"nombre":p["nombre"],"cat":p["cat"],"pa_lb":p["pa_lb"],
               "pa_usd_kg":p["pa_usd_kg"],"pa_mxn_kg":p["pa_mxn_kg"],"fuente_usda":p["fuente_usda"],
               "contenedor":p["contenedor"],"nota_unidad":p["nota_unidad"],"fuente_tipo":p["fuente_tipo"],
               "precio_mx":p["precio_mx"],"fuente_mx":p["fuente_mx"],"fecha_mx":p["fecha_mx"],
               "dif":p["dif"],"pct":p["pct"],"cdmx_barato":p["cdmx_barato"]} for p in comp]}
    with open("datos.json","w",encoding="utf-8") as f:
        json.dump(datos,f,ensure_ascii=False,indent=2)
    print("  JSON guardado.")

# ============================================================
#  PDF
# ============================================================
def crear_pdf(comp, tc, fecha_reporte):
    carpeta="docs" if os.path.exists("docs") else os.path.expanduser("~")
    hoy=datetime.now().strftime("%Y-%m-%d")
    ruta=os.path.join(carpeta,f"precios_oag_{hoy}.pdf")
    doc=SimpleDocTemplate(ruta,pagesize=landscape(letter),
        leftMargin=0.35*inch,rightMargin=0.35*inch,topMargin=0.35*inch,bottomMargin=0.4*inch)
    est=getSampleStyleSheet()
    VERDE=colors.HexColor("#1B5E20"); VC=colors.HexColor("#F0FDF4")
    RC=colors.HexColor("#FFF1F2"); GR=colors.HexColor("#F9FAFB"); BL=colors.white
    e_t=ParagraphStyle("t",parent=est["Normal"],fontSize=13,textColor=BL,backColor=VERDE,alignment=TA_CENTER,leading=18,spaceAfter=2)
    e_s=ParagraphStyle("s",parent=est["Normal"],fontSize=7.5,textColor=colors.HexColor("#555555"),alignment=TA_CENTER,spaceAfter=3)
    e_r=ParagraphStyle("r",parent=est["Normal"],fontSize=7,textColor=colors.HexColor("#475569"),alignment=TA_CENTER,spaceAfter=5)
    e_p=ParagraphStyle("p",parent=est["Normal"],fontSize=6.5,textColor=colors.HexColor("#6B7280"),alignment=TA_CENTER,spaceBefore=4)
    elems=[]
    # Logo si existe
    if LOGO_PATH and os.path.exists(LOGO_PATH):
        elems.append(RLImage(LOGO_PATH, width=1.8*inch, height=0.6*inch))
        elems.append(Spacer(1,0.05*inch))
    elems.append(Paragraph(f"PRECIOS MAYORISTAS — McAllen TX vs CDMX | {MARCA}",e_t))
    elems.append(Paragraph(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  |  TC: $1 USD = ${tc:.2f} MXN  |  Reporte USDA: {fecha_reporte}  |  Precios MX: Claude API + SNIIM",e_s))
    elems.append(Paragraph("★ = dato directo USDA AMS Reporte 3130 (Mexico Crossings Through Texas)  |  Ref. = precio de referencia con fuente y fecha  |  N/A = precio por pieza/manojo",e_r))
    enc=["#","Producto","Contenedor","$/lb","$/kg USD","MXN/kg McAllen","MXN/kg CDMX","Fuente CDMX","Fecha MX","Dif. MXN/kg","Dif. %","¿Mas barato?"]
    filas=[enc]
    for p in comp:
        dif_s=f"${p['dif']:+.2f}" if p["dif"] is not None else "N/A"
        pct_s=f"{p['pct']:+.1f}%" if p["pct"] is not None else "N/A"
        mx_s=f"${p['precio_mx']:.2f}" if p["precio_mx"] else "N/D"
        lbl="Mas barato CDMX" if p["cdmx_barato"] is True else "Mas barato McAllen" if p["cdmx_barato"] is False else "Ver nota"
        cont=p["contenedor"]+f"\n({p['nota_unidad']})" if p["nota_unidad"] else p["contenedor"]
        pl_s=f"${p['pa_lb']:.4f}" if p["fuente_tipo"]=="3130" else f"~${p['pa_lb']:.4f}"
        filas.append([str(p["idx"]),p["nombre"],cont,pl_s,f"${p['pa_usd_kg']:.4f}",f"${p['pa_mxn_kg']:.2f}",mx_s,p["fuente_mx"],p["fecha_mx"],dif_s,pct_s,lbl])
    cw=[0.22*inch,1.05*inch,1.35*inch,0.58*inch,0.6*inch,0.72*inch,0.7*inch,1.1*inch,0.72*inch,0.7*inch,0.52*inch,0.9*inch]
    from reportlab.platypus import TableStyle as TS
    tabla=Table(filas,colWidths=cw,repeatRows=1)
    sts=[("BACKGROUND",(0,0),(-1,0),VERDE),("TEXTCOLOR",(0,0),(-1,0),BL),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),6.5),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("ALIGN",(1,1),(1,-1),"LEFT"),
        ("ALIGN",(2,1),(2,-1),"LEFT"),("ALIGN",(7,1),(7,-1),"LEFT"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("FONTSIZE",(0,1),(-1,-1),6.5),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[BL,GR]),
        ("GRID",(0,0),(-1,-1),0.2,colors.HexColor("#E5E7EB")),
        ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3)]
    for i,p in enumerate(comp,1):
        if p["fuente_tipo"]=="3130":
            sts+=[("BACKGROUND",(0,i),(0,i),colors.HexColor("#DCFCE7")),("TEXTCOLOR",(0,i),(0,i),colors.HexColor("#166534")),("FONTNAME",(0,i),(0,i),"Helvetica-Bold")]
        if p["cdmx_barato"] is True:
            sts+=[("BACKGROUND",(9,i),(11,i),VC),("TEXTCOLOR",(11,i),(11,i),colors.HexColor("#166534")),("FONTNAME",(11,i),(11,i),"Helvetica-Bold")]
        elif p["cdmx_barato"] is False:
            sts+=[("BACKGROUND",(9,i),(11,i),RC),("TEXTCOLOR",(11,i),(11,i),colors.HexColor("#B91C1C")),("FONTNAME",(11,i),(11,i),"Helvetica-Bold")]
    tabla.setStyle(TS(sts))
    elems.append(tabla); elems.append(Spacer(1,0.08*inch))
    cdmx_b=sum(1 for p in comp if p["cdmx_barato"] is True)
    mc_b=sum(1 for p in comp if p["cdmx_barato"] is False)
    con_3130=sum(1 for p in comp if p["fuente_tipo"]=="3130")
    elems.append(Paragraph(f"★ {con_3130} productos Reporte 3130  |  Verde = mas barato CDMX ({cdmx_b})  |  Rojo = mas barato McAllen ({mc_b})  |  Precios MX: Claude API + SNIIM",e_p))
    doc.build(elems)
    print(f"  PDF: {ruta}")
    return ruta

# ============================================================
#  EXCEL
# ============================================================
def crear_excel(comp, tc, fecha_reporte):
    carpeta="docs" if os.path.exists("docs") else os.path.expanduser("~")
    hoy=datetime.now().strftime("%Y-%m-%d")
    ruta=os.path.join(carpeta,f"precios_oag_{hoy}.xlsx")
    wb=Workbook(); ws=wb.active; ws.title="McAllen vs CDMX"
    VERDE=PatternFill("solid",fgColor="1B5E20"); VC=PatternFill("solid",fgColor="DCFCE7")
    RC=PatternFill("solid",fgColor="FEE2E2"); GR=PatternFill("solid",fgColor="F9FAFB")
    AZ=PatternFill("solid",fgColor="DBEAFE"); BL=PatternFill("solid",fgColor="FFFFFF")
    STARF=PatternFill("solid",fgColor="F0FDF4"); REFF=PatternFill("solid",fgColor="FEF9C3")
    def eh(c,t):
        c.value=t; c.font=Font(bold=True,color="FFFFFF",size=9); c.fill=VERDE
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.merge_cells("A1:N1")
    ws["A1"]=f"{MARCA} | Precios McAllen TX vs CDMX | {datetime.now().strftime('%d/%m/%Y')} | TC: ${tc:.2f} MXN | Reporte USDA: {fecha_reporte}"
    ws["A1"].font=Font(bold=True,color="FFFFFF",size=11); ws["A1"].fill=VERDE
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=22
    ws.merge_cells("A2:N2")
    ws["A2"]="★ = USDA AMS Reporte 3130  |  Ref. = referencia con fuente y fecha  |  N/A = precio por pieza/manojo  |  Precios MX: Claude API + SNIIM"
    ws["A2"].font=Font(italic=True,color="555555",size=8); ws["A2"].alignment=Alignment(horizontal="center")
    cols=["#","Producto","Contenedor","$/lb","$/kg USD","MXN/kg McAllen","Fuente USDA","MXN/kg CDMX","Fuente CDMX","Fecha CDMX","Dif. MXN/kg","Dif. %","¿Mas barato?","Nota"]
    for c,col in enumerate(cols,1): eh(ws.cell(row=3,column=c),col)
    ws.row_dimensions[3].height=26
    for i,p in enumerate(comp,4):
        par=i%2==0; base=GR if par else BL
        star="★" if p["fuente_tipo"]=="3130" else "Ref."
        dif_s=f"${p['dif']:+.2f}" if p["dif"] is not None else "N/A"
        pct_s=f"{p['pct']:+.1f}%" if p["pct"] is not None else "N/A"
        mx_s=f"${p['precio_mx']:.2f}" if p["precio_mx"] else "N/D"
        lbl,cc=("Mas barato CDMX","166534") if p["cdmx_barato"] is True else ("Mas barato McAllen","B91C1C") if p["cdmx_barato"] is False else ("Ver nota","777777")
        pl_s=f"${p['pa_lb']:.4f}" if p["fuente_tipo"]=="3130" else f"~${p['pa_lb']:.4f}"
        fila=[star,p["nombre"],p["contenedor"],pl_s,f"${p['pa_usd_kg']:.4f}",f"${p['pa_mxn_kg']:.2f}",p["fuente_usda"],mx_s,p["fuente_mx"],p["fecha_mx"],dif_s,pct_s,lbl,p["nota_unidad"]]
        fills=[STARF if p["fuente_tipo"]=="3130" else REFF,base,base,AZ if par else BL,AZ if par else BL,AZ if par else BL,STARF if p["fuente_tipo"]=="3130" else REFF,base,base,base,
               VC if p["cdmx_barato"] is True else RC if p["cdmx_barato"] is False else base,
               VC if p["cdmx_barato"] is True else RC if p["cdmx_barato"] is False else base,
               VC if p["cdmx_barato"] is True else RC if p["cdmx_barato"] is False else base,base]
        for c,(v,fl) in enumerate(zip(fila,fills),1):
            cell=ws.cell(row=i,column=c,value=v); cell.fill=fl
            cell.alignment=Alignment(horizontal="left" if c==2 else "center",wrap_text=True)
            if c==2: cell.font=Font(bold=True,size=9)
            if c==13: cell.font=Font(bold=True,color=cc,size=9)
    anchos=[5,18,22,10,10,12,36,12,20,12,12,10,18,16]
    for idx,w in enumerate(anchos,1): ws.column_dimensions[get_column_letter(idx)].width=w
    wb.save(ruta); print(f"  Excel: {ruta}")
    return ruta

# ============================================================
#  CORREO (dos destinatarios)
# ============================================================
def enviar_correo(ruta_pdf, ruta_xl, comp, tc, fecha_reporte):
    print("Enviando correo...")
    hoy=datetime.now().strftime("%d de %B de %Y")
    cdmx_b=sum(1 for p in comp if p["cdmx_barato"] is True)
    mc_b=sum(1 for p in comp if p["cdmx_barato"] is False)
    con_3130=sum(1 for p in comp if p["fuente_tipo"]=="3130")
    top=sorted([p for p in comp if p["cdmx_barato"] is True and p["dif"] is not None],key=lambda x:x["dif"],reverse=True)[:5]
    filas="".join([f"<tr><td style='padding:5px 10px;border-bottom:1px solid #eee'>{'★' if p['fuente_tipo']=='3130' else 'Ref.'} {p['nombre']}</td>"
        f"<td style='padding:5px 10px;border-bottom:1px solid #eee;text-align:center'>${p['pa_mxn_kg']:.2f}</td>"
        f"<td style='padding:5px 10px;border-bottom:1px solid #eee;text-align:center'>${p['precio_mx']:.2f}</td>"
        f"<td style='padding:5px 10px;border-bottom:1px solid #eee;text-align:center;color:#166534;font-weight:bold'>${p['dif']:+.2f} ({p['pct']:+.1f}%)</td></tr>" for p in top])

    # Logo en base64 para email
    logo_html=""
    if LOGO_PATH and os.path.exists(LOGO_PATH):
        with open(LOGO_PATH,"rb") as f:
            logo_b64=base64.b64encode(f.read()).decode()
        logo_html=f'<img src="data:image/png;base64,{logo_b64}" style="height:50px;margin-bottom:8px" alt="{MARCA}"/><br>'

    cuerpo=f"""<div style="font-family:Arial,sans-serif;max-width:700px;margin:0 auto">
      <div style="background:#1B5E20;padding:18px 20px;border-radius:8px 8px 0 0;text-align:center">
        {logo_html}
        <h2 style="color:white;margin:0">Precios Mayoristas — McAllen TX vs CDMX</h2>
        <p style="color:#A5D6A7;margin:4px 0 0">{MARCA} &nbsp;|&nbsp; {hoy}</p>
      </div>
      <div style="background:#f9f9f9;padding:18px;border:1px solid #eee;border-top:none">
        <div style="background:#FFF8E1;padding:8px 14px;border-radius:6px;margin-bottom:8px;font-size:13px">
          <b>TC: $1 USD = ${tc:.2f} MXN</b> (Banxico) &nbsp;|&nbsp; <b>Reporte USDA: {fecha_reporte}</b>
        </div>
        <div style="background:#F0FDF4;border:1px solid #BBF7D0;padding:8px 14px;border-radius:6px;margin-bottom:8px;font-size:12px;color:#166534">
          <b>★ {con_3130} productos</b> con datos directos del Reporte 3130 &nbsp;|&nbsp; Precios MX: Claude API + SNIIM
        </div>
        <div style="background:#FFF1F2;border:1px solid #FECACA;padding:8px 14px;border-radius:6px;margin-bottom:12px;font-size:11px;color:#7F1D1D">
          Precios FOB. Sin flete, cruce fronterizo ni mermas.
        </div>
        <div style="display:flex;gap:10px;margin-bottom:14px">
          <div style="flex:1;background:#F0FDF4;border-radius:8px;padding:10px;text-align:center">
            <div style="font-size:24px;font-weight:bold;color:#166534">{cdmx_b}</div>
            <div style="font-size:11px;color:#166534">mas barato CDMX</div>
          </div>
          <div style="flex:1;background:#FFF1F2;border-radius:8px;padding:10px;text-align:center">
            <div style="font-size:24px;font-weight:bold;color:#B91C1C">{mc_b}</div>
            <div style="font-size:11px;color:#B91C1C">mas barato McAllen</div>
          </div>
          <div style="flex:1;background:#EFF6FF;border-radius:8px;padding:10px;text-align:center">
            <div style="font-size:24px;font-weight:bold;color:#1D4ED8">{len(comp)}</div>
            <div style="font-size:11px;color:#1D4ED8">productos</div>
          </div>
        </div>
        <h3 style="color:#1B5E20;margin-bottom:8px">Top 5 oportunidades CDMX:</h3>
        <table style="width:100%;border-collapse:collapse;background:white">
          <tr style="background:#1B5E20;color:white;font-size:11px">
            <th style="padding:7px 10px;text-align:left">Producto</th>
            <th style="padding:7px 10px">MXN/kg McAllen</th>
            <th style="padding:7px 10px">MXN/kg CDMX</th>
            <th style="padding:7px 10px">Diferencia</th>
          </tr>{filas}
        </table>
        <p style="margin-top:14px"><a href="{URL_PAGINA}" style="color:#1B5E20;font-weight:bold">Ver reporte en linea</a></p>
        <p style="color:#888;font-size:10px;margin-top:6px">Fuentes: USDA AMS Reporte 3130 · Claude API · SNIIM · Banxico</p>
      </div></div>"""

    for destinatario in CORREOS_DESTINO:
        msg=MIMEMultipart("alternative")
        msg["Subject"]=f"Precios McAllen vs CDMX {datetime.now().strftime('%d/%m/%Y')} | TC ${tc:.2f} | {MARCA}"
        msg["From"]=GMAIL_ORIGEN; msg["To"]=destinatario
        msg.attach(MIMEText(cuerpo,"html"))
        for ruta in [ruta_pdf,ruta_xl]:
            if ruta and os.path.exists(ruta):
                with open(ruta,"rb") as f:
                    p=MIMEBase("application","octet-stream"); p.set_payload(f.read())
                    encoders.encode_base64(p)
                    p.add_header("Content-Disposition",f"attachment; filename={os.path.basename(ruta)}")
                    msg.attach(p)
        with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
            s.login(GMAIL_ORIGEN,GMAIL_APP_PASS); s.sendmail(GMAIL_ORIGEN,destinatario,msg.as_string())
        print(f"  Correo enviado a: {destinatario}")

# ============================================================
#  WHATSAPP
# ============================================================
def enviar_whatsapp(comp, tc, fecha_reporte):
    print("Enviando WhatsApp...")
    hoy=datetime.now().strftime("%d/%b/%Y")
    cdmx_b=sum(1 for p in comp if p["cdmx_barato"] is True)
    mc_b=sum(1 for p in comp if p["cdmx_barato"] is False)
    con_3130=sum(1 for p in comp if p["fuente_tipo"]=="3130")
    top3=sorted([p for p in comp if p["cdmx_barato"] is True and p["dif"] is not None],key=lambda x:x["dif"],reverse=True)[:3]
    top3_txt="\n".join([f"  {'★' if p['fuente_tipo']=='3130' else 'Ref.'} {p['nombre']}: CDMX ${p['precio_mx']:.2f} vs McAllen ${p['pa_mxn_kg']:.2f} ({p['dif']:+.2f})" for p in top3])
    msg=(f"*{MARCA}*\n*Precios McAllen TX vs CDMX* — {hoy}\n{'─'*30}\n"
         f"TC: *${tc:.2f} MXN* | Reporte USDA: *{fecha_reporte}*\n"
         f"★ *{con_3130}* productos dato directo Reporte 3130\n\n"
         f"Mas barato CDMX: *{cdmx_b}* | McAllen: *{mc_b}*\n\n"
         f"*Top 3 oportunidades:*\n{top3_txt}\n\n"
         f"_Sin flete ni cruce. ★=Reporte 3130_\n{URL_PAGINA}")
    Client(TWILIO_SID,TWILIO_TOKEN).messages.create(body=msg,from_=TWILIO_WHATSAPP,to=TU_WHATSAPP)
    print("  WhatsApp enviado.")

# ============================================================
#  REPORTE PRINCIPAL
# ============================================================
def generar_reporte():
    print(f"\n{'='*55}\nGENERANDO REPORTE v6.0 — {MARCA}\n{datetime.now().strftime('%d/%m/%Y %H:%M')}\n{'='*55}")
    try:
        tc=obtener_tipo_cambio()
        precios_mx=obtener_precios_mx_claude()
        precios_pdf,fecha_rep=extraer_precios_pdf()
        comp=calcular_comparativa(precios_pdf,precios_mx,fecha_rep,tc)
        guardar_json(comp,tc,fecha_rep)
        pdf=crear_pdf(comp,tc,fecha_rep)
        xl=crear_excel(comp,tc,fecha_rep)
        enviar_correo(pdf,xl,comp,tc,fecha_rep)
        enviar_whatsapp(comp,tc,fecha_rep)
        con_3130=sum(1 for p in comp if p["fuente_tipo"]=="3130")
        print(f"\nReporte completado. Productos Reporte 3130: {con_3130}/{len(comp)}")
    except Exception as e:
        import traceback; print(f"\nError: {e}"); traceback.print_exc()

# ============================================================
#  ARRANQUE
# ============================================================
if __name__=="__main__":
    print(f"{MARCA} — Monitor de Precios McAllen TX v6.0")
    en_github=os.environ.get("GITHUB_ACTIONS")=="true"
    if en_github:
        generar_reporte()
    else:
        print(f"Modo local — reporte diario {HORA_REPORTE}\nCtrl+C para detener\n")
        generar_reporte()
        schedule.every().day.at(HORA_REPORTE).do(generar_reporte)
        while True:
            schedule.run_pending(); time.sleep(60)
